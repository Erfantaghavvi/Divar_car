#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Divar Car Price Scraper - Main Entry Point
اسکریپت اصلی برای جمع‌آوری قیمت خودرو از سایت‌های مختلف
"""

import os
import sys
import time
import pandas as pd
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
import time
import random
import csv
import pandas as pd
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import os
import signal
import sys
from bs4 import BeautifulSoup
import re
import concurrent.futures

# Import market price scrapers
try:
    from combined_scraper import scrape_hamrah_mechanic, scrape_z4car
    MARKET_PRICE_AVAILABLE = True
    print("✅ اسکرایپرهای قیمت بازار بهبود یافته با موفقیت بارگذاری شدند")
except ImportError as e:
    print(f"⚠️ خطا در بارگذاری اسکرایپرهای قیمت بازار: {e}")
    MARKET_PRICE_AVAILABLE = False

# Import webdriver_manager for both local and GitHub Actions
import os
from webdriver_manager.chrome import ChromeDriverManager
import subprocess  # برای صدای هشدار در مک
try:
    import winsound  # برای صدای هشدار در ویندوز
except ImportError:
    winsound = None

# متغیر سراسری برای مدیریت خروج
graceful_exit = False

def signal_handler(signum, frame):
    """مدیریت سیگنال Ctrl+C برای ذخیره داده‌ها قبل از خروج"""
    global graceful_exit
    print("\n🛑 سیگنال توقف دریافت شد (Ctrl+C)")
    print("💾 در حال ذخیره داده‌های پردازش شده...")
    graceful_exit = True
    
    # ذخیره فوری داده‌ها
    try:
        finalize_excel_save()
        print("✅ تمام داده‌ها با موفقیت ذخیره شدند")
    except Exception as e:
        print(f"❌ خطا در ذخیره داده‌ها: {e}")
    
    # بستن درایور
    try:
        if 'driver' in globals():
            driver.quit()
            print("✅ مرورگر بسته شد")
    except:
        pass
    
    print(f"🎯 تعداد کل آگهی‌های پردازش شده: {processed_count}")
    print("👋 خروج از برنامه...")
    sys.exit(0)

# ثبت signal handler
signal.signal(signal.SIGINT, signal_handler)

# ادغام با سیستم قیمت‌گذاری
try:
    from car_price_calculator import CarPriceCalculator
    PRICING_AVAILABLE = True
except ImportError as e:
    print(f"⚠️ خطا در بارگذاری سیستم قیمت‌گذاری: {e}")
    PRICING_AVAILABLE = False

try:
    from improved_car_calculator import ImprovedCarPriceCalculator
    IMPROVED_PRICING_AVAILABLE = True
    print("✅ سیستم محاسبه افت قیمت بهبود یافته بارگذاری شد")
except ImportError as e:
    print(f"⚠️ خطا در بارگذاری سیستم افت قیمت: {e}")
    IMPROVED_PRICING_AVAILABLE = False

# دیکشنری برای ذخیره قیمت‌های بازار
market_prices = {}

def fetch_market_prices(force_update=False):
    """دریافت قیمت‌های بازار از منابع مختلف با بهینه‌سازی سرعت
    
    Args:
        force_update (bool): اگر True باشد، بدون توجه به زمان فایل، داده‌ها مجدداً استخراج می‌شوند
    """
    global market_prices
    
    # بررسی وجود فایل اکسل قبلی برای جلوگیری از استخراج مجدد
    excel_file = 'combined_market_prices.xlsx'
    use_cached_data = False
    
    if not force_update and os.path.exists(excel_file):
        try:
            # بررسی تاریخ فایل - اگر کمتر از 12 ساعت از ایجاد آن گذشته باشد، از آن استفاده کن
            file_time = os.path.getmtime(excel_file)
            current_time = time.time()
            hours_diff = (current_time - file_time) / 3600
            
            # کاهش زمان به 12 ساعت برای اطمینان از به‌روز بودن داده‌ها
            if hours_diff < 12:  # اگر کمتر از 12 ساعت گذشته باشد
                print(f"📊 استفاده از قیمت‌های ذخیره شده قبلی ({hours_diff:.1f} ساعت پیش)")
                df = pd.read_excel(excel_file)
                combined_data = {}
                
                for _, row in df.iterrows():
                    name = row['نام خودرو']
                    hamrah_price = row['قیمت همراه مکانیک (تومان)']
                    z4car_price = row['قیمت زد فور (تومان)']
                    combined_data[name] = {'hamrah_mechanic': hamrah_price, 'z4car': z4car_price}
                
                market_prices = combined_data
                print(f"✅ {len(combined_data)} قیمت خودرو از فایل قبلی بارگذاری شد")
                use_cached_data = True
                return combined_data
            else:
                print(f"⚠️ فایل قیمت‌های بازار قدیمی است ({hours_diff:.1f} ساعت). در حال به‌روزرسانی...")
        except Exception as e:
            print(f"⚠️ خطا در بارگذاری فایل قبلی: {e}")
    elif force_update:
        print("🔄 درخواست به‌روزرسانی اجباری قیمت‌های بازار...")
    
    if not MARKET_PRICE_AVAILABLE:
        print("⚠️ اسکرایپرهای قیمت بازار در دسترس نیستند")
        return {}
    
    print("🔍 در حال دریافت قیمت‌های بازار...")
    
    # اجرای بهینه‌شده اسکرایپرها با تایم‌اوت کوتاه‌تر
    try:
        with concurrent.futures.ThreadPoolExecutor(max_workers=2) as executor:
            # اجرای همزمان با تایم‌اوت کاهش یافته
            hamrah_future = executor.submit(scrape_hamrah_mechanic, force_update)
            z4car_future = executor.submit(scrape_z4car, force_update)
            
            # دریافت نتایج با تایم‌اوت کاهش یافته
            hamrah_data = hamrah_future.result(timeout=120)  # کاهش به 2 دقیقه
            z4car_data = z4car_future.result(timeout=120)    # کاهش به 2 دقیقه
            
            # ادغام سریع داده‌ها - محاسبه قیمت روز از میانگین
            combined_data = {}
            
            # پردازش داده‌های همراه مکانیک
            if hamrah_data:
                for car in hamrah_data:
                    name = car.get('نام خودرو')
                    price_str = car.get('قیمت (تومان)')
                    if name and price_str:
                        # استخراج عدد از قیمت
                        price_num = int(re.sub(r'[^\d]', '', str(price_str))) if re.sub(r'[^\d]', '', str(price_str)) else 0
                        if price_num > 0:
                            combined_data[name] = {'prices': [price_num], 'count': 1}
            
            # پردازش داده‌های زد فور
            if z4car_data:
                for car in z4car_data:
                    name = car.get('نام خودرو')
                    price_str = car.get('قیمت (تومان)')
                    if name and price_str:
                        # استخراج عدد از قیمت
                        price_num = int(re.sub(r'[^\d]', '', str(price_str))) if re.sub(r'[^\d]', '', str(price_str)) else 0
                        if price_num > 0:
                            if name in combined_data:
                                combined_data[name]['prices'].append(price_num)
                                combined_data[name]['count'] += 1
                            else:
                                combined_data[name] = {'prices': [price_num], 'count': 1}
            
            # محاسبه قیمت روز از میانگین قیمت‌ها
            market_prices = {}
            for name, data in combined_data.items():
                avg_price = sum(data['prices']) / len(data['prices'])
                market_prices[name] = {'market_price': int(avg_price)}
            
            # ذخیره در فایل اکسل برای استفاده بعدی
            save_market_prices_to_excel(combined_data)
            
            print(f"✅ قیمت‌های بازار برای {len(combined_data)} خودرو دریافت شد")
            return combined_data
            
    except concurrent.futures.TimeoutError:
        print("⚠️ زمان استخراج قیمت‌ها به پایان رسید. استفاده از داده‌های موجود...")
        # تلاش برای استفاده از نتایج جزئی
        combined_data = {}
        
        try:
            if hamrah_future.done():
                hamrah_data = hamrah_future.result(timeout=1)
                for car in hamrah_data:
                    car_name = car['نام خودرو']
                    price = car['قیمت (تومان)']
                    combined_data[car_name] = {'hamrah_mechanic': price, 'z4car': None}
        except Exception as e:
            print(f"⚠️ خطا در دریافت داده‌های همراه مکانیک: {e}")
            
        try:
            if z4car_future.done():
                z4car_data = z4car_future.result(timeout=1)
                for car in z4car_data:
                    car_name = car['نام خودرو']
                    price = car['قیمت (تومان)']
                    if car_name not in combined_data:
                        combined_data[car_name] = {'hamrah_mechanic': None, 'z4car': price}
                    else:
                        combined_data[car_name]['z4car'] = price
        except Exception as e:
            print(f"⚠️ خطا در دریافت داده‌های زد فور: {e}")
            
        if combined_data:
            market_prices = combined_data
            save_market_prices_to_excel(combined_data)
            print(f"✅ قیمت‌های بازار جزئی برای {len(combined_data)} خودرو دریافت شد")
            return combined_data
        
        return {}
        
    except Exception as e:
        print(f"❌ خطا در دریافت قیمت‌های بازار: {e}")
        return {}

def save_market_prices_to_excel(data):
    """ذخیره قیمت‌های روز در فایل اکسل"""
    try:
        # تبدیل دیکشنری به دیتافریم
        rows = []
        for car_name, prices in data.items():
            row = {
                'نام خودرو': car_name,
                'قیمت روز (تومان)': prices['market_price']
            }
            rows.append(row)
        
        df = pd.DataFrame(rows)
        
        # ذخیره در فایل اکسل
        df.to_excel('combined_market_prices.xlsx', index=False)
        print("✅ قیمت‌های روز در فایل combined_market_prices.xlsx ذخیره شدند")
    except Exception as e:
        print(f"❌ خطا در ذخیره قیمت‌های روز: {e}")

def get_market_price_for_car(car_name, force_update=False):
    """دریافت قیمت بازار برای یک خودروی خاص با الگوریتم تطبیق بهینه‌شده"""
    global market_prices
    
    # کش برای جلوگیری از محاسبات تکراری
    cache_key = car_name.strip().lower()
    
    # اگر به‌روزرسانی اجباری نباشد، از کش استفاده کن
    if not force_update and hasattr(get_market_price_for_car, 'cache') and cache_key in get_market_price_for_car.cache:
        return get_market_price_for_car.cache[cache_key]
    
    if not hasattr(get_market_price_for_car, 'cache'):
        get_market_price_for_car.cache = {}
    
    # اگر به‌روزرسانی اجباری باشد یا قیمت‌های بازار موجود نباشند
    if force_update or not market_prices:
        # اگر به‌روزرسانی اجباری باشد، قیمت‌های بازار را دوباره دریافت کن
        if force_update:
            print("🔄 به‌روزرسانی اجباری قیمت‌های بازار...")
            fetch_market_prices(force_update=True)
        else:
            # اگر قیمت‌های بازار قبلاً دریافت نشده‌اند، سعی کن از فایل اکسل بخوان
            try:
                df = pd.read_excel('combined_market_prices.xlsx')
                market_prices = {}
                for _, row in df.iterrows():
                    name = row['نام خودرو']
                    market_price = row['قیمت روز (تومان)']
                    market_prices[name] = {'market_price': market_price}
                print(f"✅ قیمت‌های روز از فایل اکسل برای {len(market_prices)} خودرو بارگذاری شدند")
            except Exception as e:
                print(f"⚠️ خطا در بارگذاری قیمت‌های بازار از فایل اکسل: {e}")
                # اگر فایل وجود نداشت، قیمت‌ها را دریافت کن
                fetch_market_prices()
    
    # اگر هنوز قیمت‌های بازار در دسترس نباشند، خروج
    if not market_prices:
        result = {'market_price': None}
        get_market_price_for_car.cache[cache_key] = result
        return result
    
    # پیش‌پردازش نام خودرو برای تطبیق بهتر
    processed_car_name = car_name.lower().strip().replace('،', ' ').replace(',', ' ')
    
    # فیلتر اولیه برای کاهش تعداد مقایسه‌ها
    # ابتدا بررسی کن آیا نام دقیقاً در لیست وجود دارد
    if processed_car_name in [name.lower().strip() for name in market_prices.keys()]:
        for name in market_prices.keys():
            if name.lower().strip() == processed_car_name:
                result = market_prices[name]
                get_market_price_for_car.cache[cache_key] = result
                return result
    
    # جستجوی خودرو در قیمت‌های بازار با الگوریتم بهینه‌شده
    best_matches = []
    
    # استخراج کلمات کلیدی از نام خودرو
    car_name_words = set(processed_car_name.split())
    important_words = [w for w in car_name_words if len(w) > 2 and w not in {'و', 'با', 'در', 'از', 'به'}]
    
    # فیلتر اولیه برای کاهش تعداد مقایسه‌ها
    candidates = []
    for market_car_name in market_prices.keys():
        # بررسی سریع برای فیلتر کردن موارد نامرتبط
        market_name_lower = market_car_name.lower()
        if any(word in market_name_lower for word in important_words):
            candidates.append(market_car_name)
    
    # اگر کاندیدایی پیدا نشد، از همه خودروها استفاده کن
    if not candidates and len(market_prices) < 100:  # فقط اگر تعداد خودروها کم باشد
        candidates = list(market_prices.keys())
    
    # محاسبه امتیاز برای کاندیداها
    for market_car_name in candidates:
        score = calculate_name_match_score(car_name, market_car_name)
        if score > 0.4:  # فقط موارد با امتیاز بالاتر از آستانه را نگه دار
            best_matches.append((market_car_name, score))
    
    # مرتب‌سازی بر اساس امتیاز
    best_matches.sort(key=lambda x: x[1], reverse=True)
    
    # اگر تطابقی پیدا شد و امتیاز آن کافی است
    if best_matches and best_matches[0][1] > 0.5:
        best_match = best_matches[0][0]
        result = market_prices[best_match]
        # افزودن اطلاعات تطابق
        result['matched_name'] = best_match
        result['match_score'] = best_matches[0][1]
        # ذخیره در کش
        get_market_price_for_car.cache[cache_key] = result
        return result
    
    # اگر تطابق مناسبی پیدا نشد
    result = {'market_price': None}
    get_market_price_for_car.cache[cache_key] = result
    return result

def calculate_name_match_score(name1, name2):
    """محاسبه امتیاز تطابق بین دو نام خودرو با الگوریتم پیشرفته"""
    # پیش‌پردازش نام‌ها
    name1 = name1.lower().strip().replace('،', ' ').replace(',', ' ')
    name2 = name2.lower().strip().replace('،', ' ').replace(',', ' ')
    
    # حذف کلمات بی‌اهمیت
    stop_words = {'و', 'با', 'در', 'از', 'به', 'که', 'این', 'است', 'برای', 'یک', 'را', 'با', 'هم', 'خودرو', 'ماشین', 'اتومبیل'}
    
    # اگر نام‌ها دقیقاً یکسان باشند
    if name1 == name2:
        return 1.0
    
    # بررسی اگر یکی شامل دیگری باشد
    if name1 in name2 or name2 in name1:
        return 0.9
    
    # تقسیم به کلمات و حذف کلمات بی‌اهمیت
    words1 = [w for w in name1.split() if w not in stop_words and len(w) > 1]
    words2 = [w for w in name2.split() if w not in stop_words and len(w) > 1]
    
    # اگر پس از حذف کلمات بی‌اهمیت، کلمه‌ای باقی نماند
    if not words1 or not words2:
        return 0.0
    
    # تبدیل به مجموعه برای محاسبه اشتراک
    set_words1 = set(words1)
    set_words2 = set(words2)
    common_words = set_words1.intersection(set_words2)
    
    # محاسبه امتیاز جاکارد (نسبت اشتراک به اجتماع)
    jaccard_score = len(common_words) / len(set_words1.union(set_words2)) if set_words1 or set_words2 else 0
    
    # محاسبه امتیاز تطابق دقیق کلمات (با وزن بیشتر)
    exact_match_score = len(common_words) / min(len(set_words1), len(set_words2)) if common_words else 0
    
    # بررسی تطابق برند و مدل (کلمات اول و دوم معمولاً برند و مدل هستند)
    brand_model_score = 0
    if len(words1) >= 1 and len(words2) >= 1:
        # بررسی تطابق کلمه اول (معمولاً برند)
        if words1[0] == words2[0]:
            brand_model_score += 0.5
        # بررسی تطابق کلمه دوم (معمولاً مدل) اگر وجود داشته باشد
        if len(words1) >= 2 and len(words2) >= 2 and words1[1] == words2[1]:
            brand_model_score += 0.3
    
    # بررسی تطابق جزئی کلمات (برای مواردی که کلمات مشابه اما نه دقیقاً یکسان هستند)
    partial_match_score = 0
    for w1 in words1:
        for w2 in words2:
            if w1 != w2 and (w1 in w2 or w2 in w1) and min(len(w1), len(w2)) > 2:
                # امتیاز بر اساس نسبت طول کلمه کوتاه‌تر به بلندتر
                partial_match_score += min(len(w1), len(w2)) / max(len(w1), len(w2))
    
    # نرمال‌سازی امتیاز تطابق جزئی
    max_possible_partial = max(len(words1), len(words2))
    if max_possible_partial > 0:
        partial_match_score = min(partial_match_score / max_possible_partial, 1.0)
    
    # ترکیب امتیازها با وزن‌های مختلف
    final_score = (jaccard_score * 0.3) + (exact_match_score * 0.3) + (brand_model_score * 0.3) + (partial_match_score * 0.1)
    
    # محدود کردن امتیاز نهایی بین 0 و 1
    return min(max(final_score, 0.0), 1.0)


##driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))

# مسیر پروفایل کاربری منحصر به فرد

import os
profile_path = os.path.expanduser(f'~/chrome_divar_profile_{int(time.time())}')

# تنظیمات بهینه Chrome برای macOS ARM64 و حل مشکل timeout
chrome_options = Options()

# تنظیمات اساسی برای macOS ARM64
chrome_options.add_argument('--no-sandbox')
chrome_options.add_argument('--disable-dev-shm-usage')

# حل مشکل timeout در macOS ARM64 - تنظیمات پیشرفته
chrome_options.add_argument('--disable-features=VizDisplayCompositor,VizServiceDisplayCompositor')
chrome_options.add_argument('--disable-gpu-sandbox')
chrome_options.add_argument('--disable-software-rasterizer')
chrome_options.add_argument('--disable-background-timer-throttling')
chrome_options.add_argument('--disable-backgrounding-occluded-windows')
chrome_options.add_argument('--disable-renderer-backgrounding')
chrome_options.add_argument('--disable-field-trial-config')
chrome_options.add_argument('--disable-back-forward-cache')
chrome_options.add_argument('--disable-ipc-flooding-protection')
chrome_options.add_argument('--max_old_space_size=4096')
chrome_options.add_argument('--js-flags=--max-old-space-size=4096')

# تنظیمات اضافی برای حل مشکل timeout
chrome_options.add_argument('--disable-dev-tools')
chrome_options.add_argument('--disable-logging')
chrome_options.add_argument('--disable-gpu-process-crash-limit')
chrome_options.add_argument('--disable-crash-reporter')
chrome_options.add_argument('--no-crash-upload')
chrome_options.add_argument('--disable-in-process-stack-traces')
chrome_options.add_argument('--disable-logging-redirect')
chrome_options.add_argument('--log-level=3')
chrome_options.add_argument('--silent')
chrome_options.add_argument('--disable-gpu-process-for-dx12-vulkan-info-collection')
chrome_options.add_argument('--disable-gpu-process-for-dx12-vulkan-info-collection')
chrome_options.add_argument('--single-process')  # اجرا در یک پروسه برای کاهش مشکلات IPC

# تنظیمات بهینه‌سازی سرعت
chrome_options.add_argument('--disable-gpu')
chrome_options.add_argument('--disable-extensions')
chrome_options.add_argument('--disable-plugins')
chrome_options.add_argument('--disable-infobars')
chrome_options.add_argument('--disable-notifications')
chrome_options.add_argument('--disable-popup-blocking')
chrome_options.add_argument('--disable-save-password-bubble')
chrome_options.add_argument('--disable-translate')
chrome_options.add_argument('--disable-web-security')
chrome_options.add_argument('--disable-features=TranslateUI,BlinkGenPropertyTrees')
chrome_options.add_argument('--disable-site-isolation-trials')
chrome_options.add_argument('--ignore-certificate-errors')
chrome_options.add_argument('--disable-blink-features=AutomationControlled')
chrome_options.add_experimental_option('excludeSwitches', ['enable-automation'])
chrome_options.add_experimental_option('useAutomationExtension', False)

# تنظیمات حافظه
chrome_options.add_argument('--aggressive-cache-discard')
chrome_options.add_argument('--disable-application-cache')
chrome_options.add_argument('--disable-cache')
chrome_options.add_argument('--disable-offline-load-stale-cache')
chrome_options.add_argument('--disk-cache-size=0')

# تنظیمات پردازش
chrome_options.add_argument('--disable-background-networking')
chrome_options.add_argument('--disable-background-timer-throttling')
chrome_options.add_argument('--disable-backgrounding-occluded-windows')
chrome_options.add_argument('--disable-breakpad')
chrome_options.add_argument('--disable-component-extensions-with-background-pages')
chrome_options.add_argument('--disable-default-apps')
chrome_options.add_argument('--disable-hang-monitor')
chrome_options.add_argument('--disable-ipc-flooding-protection')
chrome_options.add_argument('--disable-prompt-on-repost')
chrome_options.add_argument('--disable-renderer-backgrounding')
chrome_options.add_argument('--disable-sync')
chrome_options.add_argument('--disable-domain-reliability')
chrome_options.add_argument('--disable-client-side-phishing-detection')
chrome_options.add_argument('--disable-features=UserAgentClientHint')
chrome_options.add_argument('--metrics-recording-only')
chrome_options.add_argument('--no-default-browser-check')
chrome_options.add_argument('--no-first-run')
chrome_options.add_argument('--no-pings')
chrome_options.add_argument('--password-store=basic')
chrome_options.add_argument('--use-mock-keychain')
chrome_options.add_argument('--process-per-tab')
chrome_options.add_argument('--enable-low-end-device-mode')

def create_driver():
    """ایجاد driver بهینه‌سازی شده برای macOS ARM64"""
    try:
        print("🚀 راه‌اندازی Chrome driver...")
        
        # استفاده از webdriver-manager یا chromedriver سیستم
        try:
            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=chrome_options)
        except Exception as e:
            print(f"⚠️ خطا در webdriver-manager: {e}")
            print("🔄 تلاش برای استفاده از chromedriver سیستم...")
            driver = webdriver.Chrome(options=chrome_options)
        
        # تنظیمات timeout بهینه برای macOS ARM64
        driver.set_page_load_timeout(60)  # افزایش بیشتر timeout
        driver.implicitly_wait(10)  # افزایش implicit wait
        driver.set_script_timeout(60)  # تنظیم script timeout
        
        # مخفی کردن ماهیت وب‌درایور
        driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        
        # تنظیم اندازه پنجره برای بهینه‌سازی مصرف منابع
        driver.set_window_size(1366, 768)
        
        # تست اولیه برای اطمینان از عملکرد driver
        try:
            driver.get("data:text/html,<html><body><h1>Test</h1></body></html>")
            time.sleep(1)
        except Exception as test_error:
            print(f"⚠️ خطا در تست اولیه driver: {test_error}")
            # ادامه می‌دهیم چون ممکن است در استفاده واقعی کار کند
        
        print("✅ Chrome driver راه‌اندازی شد")
        return driver
    except Exception as e:
        print(f"❌ خطا در ایجاد driver: {e}")
        return None

def safe_driver_operation(operation_func, *args, max_retries=3, **kwargs):
    """اجرای ایمن عملیات driver با retry برای مشکلات timeout"""
    for attempt in range(max_retries):
        try:
            return operation_func(*args, **kwargs)
        except Exception as e:
            error_str = str(e).lower()
            if 'timeout' in error_str and attempt < max_retries - 1:
                print(f"⚠️ Timeout در تلاش {attempt + 1}, تلاش مجدد...")
                time.sleep(5)  # انتظار بیشتر قبل از تلاش مجدد
                continue
            else:
                raise e
    return None

def safe_page_load(driver, url, max_retries=5):
    """بارگذاری ایمن صفحه با retry برای مشکلات timeout"""
    for attempt in range(max_retries):
        try:
            print(f"🔄 تلاش {attempt + 1} برای بارگذاری: {url}")
            driver.get(url)
            print("✅ صفحه با موفقیت بارگذاری شد")
            return True
        except Exception as e:
            error_str = str(e).lower()
            if 'timeout' in error_str and attempt < max_retries - 1:
                print(f"⚠️ Timeout در تلاش {attempt + 1}, تلاش مجدد در 10 ثانیه...")
                time.sleep(10)  # انتظار بیشتر برای timeout
                continue
            else:
                print(f"❌ خطا در بارگذاری صفحه: {e}")
                if attempt == max_retries - 1:
                    raise e
    return False



def restart_driver():
    """راه‌اندازی مجدد driver در صورت crash بدون از دست دادن موقعیت فعلی"""
    global driver
    current_url = None
    
    # ذخیره URL فعلی قبل از restart
    try:
        if 'driver' in globals() and driver:
            current_url = driver.current_url
            print(f"💾 ذخیره موقعیت فعلی: {current_url}")
    except:
        current_url = "https://divar.ir/s/iran/car"  # fallback به صفحه اصلی
    
    # بستن درایور قبلی به صورت ایمن
    try:
        if 'driver' in globals() and driver:
            driver.quit()
            print("🔒 درایور قبلی با موفقیت بسته شد")
    except Exception as e:
        print(f"⚠️ خطا در بستن درایور قبلی: {e}")
    
    print("🔄 راه‌اندازی مجدد Chrome driver...")
    driver = create_driver()
    if driver:
        # بازگشت به موقعیت قبلی با مدیریت خطا
        try:
            if current_url and current_url != "data:,":
                print(f"🔄 بازگشت به موقعیت قبلی: {current_url}")
                safe_page_load(driver, current_url)
                # کاهش زمان انتظار تصادفی برای افزایش سرعت
                time.sleep(random.uniform(1.5, 3))
            else:
                print("🏠 رفتن به صفحه اصلی")
                safe_page_load(driver, "https://divar.ir/s/iran/car")
                time.sleep(random.uniform(1, 2))
        except Exception as e:
            print(f"⚠️ خطا در بازگشت به موقعیت قبلی: {e}")
            print("🏠 تلاش مجدد برای رفتن به صفحه اصلی")
            try:
                safe_page_load(driver, "https://divar.ir/s/iran/car")
                time.sleep(1)
            except:
                pass
        
        print("✅ Driver با موفقیت راه‌اندازی شد و موقعیت بازیابی شد")
        return True
    else:
        print("❌ خطا در راه‌اندازی مجدد driver")
        return False

# راه‌اندازی اولیه درایور
# بارگذاری داده‌های موجود قبل از شروع
# load_existing_data() - این تابع در ادامه کد تعریف شده و در حلقه اصلی فراخوانی می‌شود

driver = create_driver()
if not driver:
    exit(1)

# اسکریپت‌های فوق پیشرفته برای مخفی کردن کامل ماهیت ربات
# حذف اسکریپت stealth برای جلوگیری از crash

# تنظیم اندازه پنجره ثابت
driver.set_window_size(1366, 768)

# رفتن به صفحه خودرو با retry mechanism
if not safe_page_load(driver, 'https://divar.ir/s/iran/car'):
    print("❌ نتوانستیم به صفحه اصلی برویم، خروج از برنامه")
    exit(1)
time.sleep(2)


def append_to_csv_old(a, b, filename="data.csv"):
    # Open the CSV file in append mode
    with open(filename, mode='a', newline='') as file:
        writer = csv.writer(file)
        # Append the values of a and b as a new row
        writer.writerow([a, b])


def append_to_csv(a, b, filename="data.csv"):
    # Open the CSV file in append mode with utf-8 encoding
    with open(filename, mode='a', newline='', encoding='utf-8') as file:
        writer = csv.writer(file)
        # Append the values of a and b as a new row
        writer.writerow([a, b])


# متغیرهای سراسری برای شمارش و ذخیره‌سازی
record_count = 0
data_batch = []
BATCH_SIZE = 50  # افزایش اندازه batch برای عملکرد بهتر

def load_existing_data():
    """بارگذاری داده‌های موجود (اختیاری)"""
    global record_count
    
    # شروع جدید بدون وابستگی به فایل خاص
    print("🚀 شروع جلسه جدید استخراج آگهی‌ها")
    record_count = 0
    return pd.DataFrame()

def save_to_excel(ad_data, batch_size=BATCH_SIZE):
    """ذخیره اطلاعات آگهی با ذخیره دوره‌ای"""
    global record_count, data_batch
    
    try:
        # اضافه کردن داده جدید به batch
        data_batch.append(ad_data)
        record_count += 1
        
        # نمایش پیشرفت
        if record_count % 100 == 0:
            print(f"✅ {record_count} آگهی پردازش شده...")
        elif record_count % 50 == 0:
            print(f"📊 {record_count} آگهی جمع‌آوری شده...")
        
        # ذخیره دوره‌ای هر BATCH_SIZE آگهی
        if len(data_batch) >= batch_size:
            save_batch_to_improved_file()
            
        return True
            
    except Exception as e:
        print(f"خطا در ذخیره داده: {e}")
        # در صورت خطا، در فایل CSV ذخیره کنیم
        try:
            df_new = pd.DataFrame([ad_data])
            from datetime import datetime
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            csv_filename = f"divar_ads_backup_{timestamp}.csv"
            df_new.to_csv(csv_filename, mode='a', header=not os.path.exists(csv_filename), index=False, encoding='utf-8')
            print(f"✅ اطلاعات در فایل پشتیبان CSV ذخیره شد")
        except:
            print("❌ خطا در ذخیره اطلاعات")
        return True

def save_batch_to_improved_file():
    """ذخیره batch فعلی در فایل بهبود یافته"""
    global data_batch
    
    if not data_batch:
        return
        
    try:
        # ایجاد نام فایل با timestamp
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"improved_divar_ads_{timestamp}.xlsx"
        
        # تبدیل batch به DataFrame
        df_new = pd.DataFrame(data_batch)
        
        # ذخیره با رنگ‌بندی
        save_with_coloring(df_new, filename)
        
        print(f"💾 {len(df_new)} رکورد در فایل {filename} ذخیره شد")
        
        # پاک کردن batch
        data_batch = []
        
    except Exception as e:
        print(f"❌ خطا در ذخیره batch: {e}")
        # ذخیره اضطراری
        try:
            from datetime import datetime
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            emergency_file = f"divar_ads_emergency_{timestamp}.xlsx"
            df_emergency = pd.DataFrame(data_batch)
            df_emergency.to_excel(emergency_file, index=False, engine='openpyxl')
            print(f"🚨 ذخیره اضطراری در {emergency_file}")
            data_batch = []
        except Exception as e2:
            print(f"❌ خطا در ذخیره اضطراری: {e2}")

def save_with_coloring(df, filename):
    """ذخیره DataFrame با رنگ‌بندی سبز برای آگهی‌های فوری"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # اطمینان از وجود دایرکتوری
        os.makedirs(os.path.dirname(os.path.abspath(filename)), exist_ok=True)
        
        # اطمینان از اینکه DataFrame خالی نیست
        if df.empty:
            print("⚠️ DataFrame خالی است - فایل اکسل ایجاد نمی‌شود")
            return
            
        # ایجاد workbook
        wb = Workbook()
        ws = wb.active
        
        # اضافه کردن داده‌ها
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # رنگ‌بندی سبز برای آگهی‌های فوری
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        
        # پیدا کردن ستون رنگ‌بندی
        color_col_index = None
        for idx, col in enumerate(df.columns, 1):
            if col == 'رنگ‌بندی':
                color_col_index = idx
                break
        
        if color_col_index:
            # رنگ‌بندی ردیف‌های سبز
            urgent_count = 0
            for row_idx in range(2, len(df) + 2):  # شروع از ردیف 2 (بعد از header)
                cell_value = ws.cell(row=row_idx, column=color_col_index).value
                if cell_value == 'سبز':
                    urgent_count += 1
                    for col_idx in range(1, len(df.columns) + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = green_fill
            
            if urgent_count > 0:
                print(f"✅ {urgent_count} آگهی فوری با رنگ سبز مشخص شدند")
        
        # تنظیم عرض ستون‌ها برای خوانایی بهتر
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = min(adjusted_width, 50)  # حداکثر عرض 50
        
        # ذخیره فایل با مدیریت خطا
        try:
            wb.save(filename)
            print(f"✅ فایل اکسل با موفقیت در {filename} ذخیره شد")
        except PermissionError:
            # تلاش برای ذخیره با نام متفاوت در صورت باز بودن فایل
            from datetime import datetime
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            alt_filename = f"{os.path.splitext(filename)[0]}_{timestamp}{os.path.splitext(filename)[1]}"
            wb.save(alt_filename)
            print(f"⚠️ فایل اصلی باز است - ذخیره در {alt_filename}")
    except Exception as e:
        print(f"❌ خطا در ذخیره فایل اکسل: {e}")
        # تلاش برای ذخیره ساده بدون رنگ‌بندی
        try:
            df.to_excel(f"{os.path.splitext(filename)[0]}_simple.xlsx", index=False, engine='openpyxl')
            print("🔄 فایل اکسل ساده (بدون رنگ‌بندی) ذخیره شد")
        except Exception as e2:
            print(f"❌ خطا در ذخیره ساده: {e2}")
        
    except Exception as e:
        print(f"خطا در رنگ‌بندی: {e} - ذخیره ساده")
        # ذخیره ساده بدون رنگ‌بندی
        df.to_excel(filename, index=False, engine='openpyxl')

def finalize_excel_save():
    """ذخیره نهایی تمام داده‌ها در فایل اکسل بهبود یافته مشابه improved_divar_ads"""
    global record_count, data_batch
    
    try:
        # ایجاد timestamp برای نام فایل
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"improved_divar_ads_{timestamp}.xlsx"
        
        # استفاده از داده‌های batch فعلی
        all_data = data_batch
        print(f"\n💾 آماده‌سازی {len(all_data)} آگهی برای ذخیره نهایی...")
        
        # اگر قیمت‌های بازار هنوز دریافت نشده‌اند، آنها را دریافت کن
        if MARKET_PRICE_AVAILABLE and not market_prices:
            print("🔍 دریافت قیمت‌های بازار برای ادغام با داده‌های دیوار...")
            fetch_market_prices(force_update=True)
        
        # ایجاد DataFrame با ستون‌های مرتب شده مشابه improved_divar_ads
        # (آگهی‌ها قبلاً پردازش شده‌اند)
        df = pd.DataFrame(all_data)
        
        try:
            # افزودن قیمت‌های بازار به داده‌ها
            if MARKET_PRICE_AVAILABLE:
                print("🔄 در حال ادغام قیمت‌های بازار با داده‌های دیوار...")
                
                # اگر قیمت‌های بازار هنوز دریافت نشده‌اند یا نیاز به به‌روزرسانی دارند، آنها را دریافت کن
                if not market_prices:
                    print("🔍 دریافت قیمت‌های بازار برای ادغام با داده‌های دیوار...")
                    fetch_market_prices(force_update=True)
                
                # قیمت روز قبلاً در process_ad_with_pricing محاسبه شده است
                print("قیمت‌های روز در مرحله پردازش آگهی‌ها محاسبه شده‌اند")
        except Exception as e:
            print(f"❌ خطا در افزودن قیمت‌های بازار: {e}")
        
        print(f"✅ قیمت‌های بازار برای {len(df)} خودرو اضافه شد")
        
        # مرتب‌سازی ستون‌ها مطابق فرمت مطلوب (بدون ستون‌های همراه مکانیک و زد فور)
        desired_columns = [
            'عنوان آگهی', 'برند و تیپ', 'نام خودرو', 'سال', 'کیلومتر', 'رنگ',
            'قیمت آگهی (تومان)', 'قیمت روز (تومان)', 'قیمت تخمینی (تومان)', 'منبع قیمت',
            'وضعیت موتور', 'وضعیت شاسی', 'وضعیت بدنه', 'مشکلات تشخیص داده شده',
            'مشکلات موتور/گیربکس', 'فروش فوری', 'افت کارکرد', 'افت سن خودرو', 
            'افت مشکلات', 'درصد افت کل', 'توضیحات', 'گیربکس', 'بیمه شخص ثالث',
            'معاوضه', 'دسته بندی خودرو', 'شماره تلفن', 'زمان و مکان', 'لینک آگهی',
            'رنگ‌بندی', 'تاریخ ذخیره'
        ]
        
        # اضافه کردن تاریخ ذخیره
        df['تاریخ ذخیره'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # مرتب‌سازی ستون‌ها
        available_columns = [col for col in desired_columns if col in df.columns]
        remaining_columns = [col for col in df.columns if col not in desired_columns]
        final_columns = available_columns + remaining_columns
        df = df[final_columns]
        
        # ذخیره با رنگ‌بندی
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        wb = Workbook()
        ws = wb.active
        ws.title = "آگهی‌های دیوار بهبود یافته"
        
        # اضافه کردن داده‌ها
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # رنگ‌بندی سبز برای آگهی‌های فوری
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        
        # پیدا کردن ستون رنگ‌بندی
        color_col_index = None
        for idx, col in enumerate(df.columns, 1):
            if col == 'رنگ‌بندی':
                color_col_index = idx
                break
        
        urgent_count = 0
        if color_col_index:
            # رنگ‌بندی ردیف‌های سبز
            for row_idx in range(2, len(df) + 2):  # شروع از ردیف 2 (بعد از header)
                cell_value = ws.cell(row=row_idx, column=color_col_index).value
                if cell_value == 'سبز':
                    urgent_count += 1
                    for col_idx in range(1, len(df.columns) + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = green_fill
        
        # ذخیره فایل بهبود یافته با رنگ‌بندی
        try:
            # تنظیم عرض ستون‌ها برای خوانایی بهتر
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = min(adjusted_width, 50)  # حداکثر عرض 50
            
            wb.save(filename)
            print(f"✅ فایل بهبود یافته ذخیره شد: {filename}")
        except Exception as e:
            print(f"⚠️ خطا در ذخیره فایل بهبود یافته: {e}")
            # تلاش برای ذخیره با نام متفاوت
            try:
                alt_filename = f"improved_divar_ads_{timestamp}_alt.xlsx"
                wb.save(alt_filename)
                print(f"✅ فایل بهبود یافته با نام جایگزین ذخیره شد: {alt_filename}")
            except Exception as e2:
                print(f"❌ خطا در ذخیره جایگزین: {e2}")
        
        # همچنین داده‌ها را در فایل اصلی نیز ذخیره کنیم با استفاده از تابع save_with_coloring
        try:
            save_with_coloring(df, MAIN_EXCEL_FILE)
            print(f"✅ فایل اصلی به‌روزرسانی شد: {MAIN_EXCEL_FILE}")
        except Exception as e:
            print(f"⚠️ خطا در به‌روزرسانی فایل اصلی: {e}")
            # تلاش برای ذخیره ساده
            try:
                df.to_excel(MAIN_EXCEL_FILE, index=False, engine='openpyxl')
                print(f"✅ فایل اصلی با روش ساده به‌روزرسانی شد")
            except Exception as e2:
                print(f"❌ خطا در به‌روزرسانی ساده فایل اصلی: {e2}")
        
        print(f"📊 آمار نهایی:")
        print(f"   - تعداد کل آگهی‌ها: {len(df)}")
        print(f"   - آگهی‌های فوری (سبز): {urgent_count}")
        print(f"   - تعداد ستون‌ها: {len(df.columns)}")
        
        data_batch = []
        
    except Exception as e:
        print(f"❌ خطا در ذخیره نهایی: {e}")
        # در صورت خطا، ذخیره ساده
        try:
            df = pd.DataFrame(data_batch)
            df.to_excel(f"divar_ads_backup_{timestamp}.xlsx", index=False, engine='openpyxl')
            print(f"🚨 ذخیره اضطراری انجام شد")
            data_batch = []
        except Exception as e2:
            print(f"❌ خطا در ذخیره اضطراری: {e2}")


# تابع کپچا حذف شد - فقط تایم‌های انسانی استفاده می‌شود


animation_script_toblue = """
        let element = arguments[0];

        // Store the original font size and color
        let originalFontSize = window.getComputedStyle(element).fontSize;
        let originalColor = window.getComputedStyle(element).color;

        // حذف انیمیشن برای سرعت بیشتر - تغییر فوری
        element.style.fontSize = 'calc(' + originalFontSize + ' + 1rem)';
        element.style.color = 'white';
        
        // بازگشت سریع
        setTimeout(() => {
            element.style.fontSize = originalFontSize;
            element.style.color = originalColor;
        }, 100);  // کاهش زمان به 100ms
"""

animation_script = """
        let element = arguments[0];

        // Store the original background color
        let originalBackground = window.getComputedStyle(element).backgroundColor;

        // Get element's bounding rectangle to calculate position
        let rect = element.getBoundingClientRect();

        // حذف انیمیشن برای سرعت بیشتر - تغییر فوری
        element.style.transform = 'scale(1.2)';
        element.style.backgroundColor = 'rgba(255, 0, 0, 0.8)';
        
        // بازگشت سریع
        setTimeout(() => {
            element.style.transform = 'scale(1)';
            element.style.backgroundColor = originalBackground;
        }, 100);  // کاهش زمان به 100ms 
    """


def Find_Elements_By_XPATH(d, x):
    WebDriverWait(d, 3).until(
        EC.presence_of_all_elements_located((By.XPATH, x))
    )
    return d.find_elements(By.XPATH, x)


def Find_Element_By_XPATH(d, x):
    WebDriverWait(d, 2).until(
        EC.presence_of_element_located((By.XPATH, x))
    )
    return d.find_element(By.XPATH, x)


# تابع حرکت موس حذف شد برای ساده‌سازی

def human_typing(element, text):
    """تایپ انسانی با سرعت متغیر و اشتباهات"""
    element.clear()
    time.sleep(random.uniform(0.5, 1.5))  # تفکر قبل از تایپ
    
    for i, char in enumerate(text):
        # گاهی اشتباه تایپ کن و اصلاح کن
        if random.random() < 0.05 and i > 0:  # 5% احتمال اشتباه
            wrong_char = random.choice('abcdefghijklmnopqrstuvwxyz')
            element.send_keys(wrong_char)
            time.sleep(random.uniform(0.1, 0.3))
            element.send_keys('\b')  # پاک کردن
            time.sleep(random.uniform(0.1, 0.2))
        
        element.send_keys(char)
        # سرعت تایپ متغیر
        if char == ' ':
            time.sleep(random.uniform(0.2, 0.5))  # مکث بیشتر بعد از فاصله
        else:
            time.sleep(random.uniform(0.1, 0.4))
    
    time.sleep(random.uniform(0.3, 0.8))  # مکث بعد از تایپ


# حالا مستقیماً در صفحه خودروها هستیم، نیازی به سرچ نیست
# divar_search = Find_Element_By_XPATH(driver, '//*[@id="app"]/header/nav/div/div[2]/div/div/div[1]/form/input')
# driver.execute_script(animation_script, divar_search)
# human_typing(divar_search, "النترا")
# time.sleep(0.2)
# divar_search.send_keys(Keys.RETURN)


# time.sleep(10000)

def check_for_critical_bot_detection():
    """تشخیص مشکلات جدی که نیاز به مداخله دارند"""
    try:
        # بررسی کپچاهای مختلف
        captcha_selectors = [
            "iframe[src*='recaptcha']",
            ".g-recaptcha",
            "iframe[src*='hcaptcha']",
            "#arc-checkbox",  # ARCaptcha
            "[id*='captcha']",
            "[class*='captcha']",
            "input[name='arcaptcha-token']"
        ]
        
        for selector in captcha_selectors:
            elements = driver.find_elements(By.CSS_SELECTOR, selector)
            if elements and any(elem.is_displayed() for elem in elements):
                return "CAPTCHA_DETECTED"
        
        # بررسی متن‌های مربوط به کپچا
        captcha_texts = [
            "من ربات نیستم",
            "i'm not a robot",
            "verify you are human",
            "arcaptcha",
            "recaptcha"
        ]
        
        page_text = driver.page_source.lower()
        for text in captcha_texts:
            if text.lower() in page_text:
                return f"CAPTCHA_TEXT: {text}"
        
        # بررسی پیام‌های مسدودیت جدی
        critical_blocking_texts = [
            "تایید هویت",
            "مسدود شده",
            "blocked permanently",
            "403 forbidden",
            "access denied",
            "too many requests"
        ]
        
        for text in critical_blocking_texts:
            if text.lower() in page_text:
                return f"CRITICAL_BLOCK: {text}"
        
        # بررسی ریدایرکت به صفحات غیرمرتبط
        current_url = driver.current_url.lower()
        if "divar.ir" not in current_url and "google.com" not in current_url and "arcaptcha" not in current_url:
            return f"REDIRECT: {current_url}"
            
        return None
        
    except Exception as e:
        return None

def handle_verification_issue():
    """مدیریت مشکلات تایید هویت"""
    try:
        print("🔄 تلاش برای بازگشت به صفحه اصلی...")
        # بازگشت به صفحه اصلی خودروها
        safe_page_load(driver, 'https://divar.ir/s/iran/car?map_interaction=search_this_area_disabled')
        time.sleep(random.uniform(3, 6))

        return True
    except Exception as e:

        return False

def alert_user_critical(message):
    """هشدار فقط برای مشکلات جدی"""
    print("\n" + "="*60)
    print("🚨 هشدار جدی: مشکل تشخیص داده شد! 🚨")
    print(f"📋 مشکل: {message}")
    
    # پخش صدای هشدار
    try:
        for _ in range(2):  # دو بار صدا
            subprocess.run(["afplay", "/System/Library/Sounds/Sosumi.aiff"], check=False)
            time.sleep(0.3)
    except:
        try:
            if winsound:
                for _ in range(2):
                    winsound.Beep(1500, 500)
                    time.sleep(0.2)
        except:
            print("🔔" * 10 + " توجه! " + "🔔" * 10)
    
    # ادامه خودکار بدون انتظار برای ورودی کاربر
    print("⏳ صبر 2 ثانیه و سپس ادامه خودکار...")
    time.sleep(2)  # صبر 2 ثانیه
    print("▶️ ادامه کار ربات...")
    
    print("="*60 + "\n")

#def extract_phone_number(driver):
    # بررسی فقط مشکلات جدی
    critical_issue = check_for_critical_bot_detection()
    if critical_issue:
        alert_user_critical(critical_issue)

    WebDriverWait(driver, 2).until(
        EC.presence_of_all_elements_located((By.TAG_NAME, 'a'))
    )
    links = driver.find_elements(By.TAG_NAME, 'a')

    phone_number = "0"



def extract_ad_details():
    """استخراج کامل اطلاعات آگهی از دیوار"""
    ad_data = {
        'عنوان آگهی': '',
        'برند و تیپ': '',
        'نام خودرو': '',
        'سال': '',
        'کیلومتر': '',
        'رنگ': '',
        'قیمت آگهی (تومان)': '',
        'قیمت روز (تومان)': '',
        'منبع قیمت': 'دیوار',
        'وضعیت موتور': '',
        'وضعیت شاسی': '',
        'وضعیت بدنه': '',
        'مشکلات تشخیص داده شده': '',
        'افت کارکرد': '',
        'افت سن خودرو': '',
        'افت مشکلات': '',
        'درصد افت کل': '',
        'قیمت تخمینی (تومان)': '',
        'توضیحات': '',
        'گیربکس': '',
        'بیمه شخص ثالث': '',
        'معاوضه': '',
        'دسته بندی خودرو': 'خودرو',
        'شماره تلفن': '',
        'زمان و مکان': '',
        'لینک آگهی': driver.current_url
    }
    
    # عنوان آگهی
    try:
        title_elem = driver.find_element(By.TAG_NAME, "h1")
        ad_data['عنوان آگهی'] = title_elem.text.strip()
    except:
        try:
            title_elem = driver.find_element(By.CSS_SELECTOR, ".kt-page-title__title")
            ad_data['عنوان آگهی'] = title_elem.text.strip()
        except:
            pass
    
    # استخراج سال با CSS Selector جدید
    try:
        year_element = driver.find_element(By.CSS_SELECTOR, "#app > div.container--has-footer-d86a9.kt-container > div > main > article > div > div.kt-col-5 > section:nth-child(1) > div.post-page__section--padded > table > tbody > tr > td:nth-child(2)")
        ad_data['سال'] = year_element.text.strip()
        print(f"✅ سال از CSS Selector جدید: {ad_data['سال']}")
    except Exception as e:
        print(f"⚠️ خطا در استخراج سال از CSS Selector جدید: {e}")
        # fallback به روش قبلی
        try:
            cells = driver.find_elements(By.CSS_SELECTOR, ".kt-group-row-item__value")
            if len(cells) >= 2:
                ad_data['سال'] = cells[1].text.strip()
        except:
            pass
    
    # استخراج کیلومتر با CSS Selector جدید
    try:
        km_element = driver.find_element(By.CSS_SELECTOR, "#app > div.container--has-footer-d86a9.kt-container > div > main > article > div > div.kt-col-5 > section:nth-child(1) > div.post-page__section--padded > table > tbody > tr > td:nth-child(1)")
        ad_data['کیلومتر'] = km_element.text.strip()
        print(f"✅ کیلومتر از CSS Selector جدید: {ad_data['کیلومتر']}")
    except Exception as e:
        print(f"⚠️ خطا در استخراج کیلومتر از CSS Selector جدید: {e}")
        # fallback به روش قبلی
        try:
            cells = driver.find_elements(By.CSS_SELECTOR, ".kt-group-row-item__value")
            if len(cells) >= 1:
                ad_data['کیلومتر'] = cells[0].text.strip()
        except:
            pass
    
    # استخراج رنگ با CSS Selector جدید
    try:
        color_element = driver.find_element(By.CSS_SELECTOR, "#app > div.container--has-footer-d86a9.kt-container > div > main > article > div > div.kt-col-5 > section:nth-child(1) > div.post-page__section--padded > table > tbody > tr > td:nth-child(3)")
        ad_data['رنگ'] = color_element.text.strip()
        print(f"✅ رنگ از CSS Selector جدید: {ad_data['رنگ']}")
    except Exception as e:
        print(f"⚠️ خطا در استخراج رنگ از CSS Selector جدید: {e}")
        # fallback به روش قبلی
        try:
            cells = driver.find_elements(By.CSS_SELECTOR, ".kt-group-row-item__value")
            if len(cells) >= 3:
                ad_data['رنگ'] = cells[2].text.strip()
        except:
            pass
            
    # قیمت آگهی با CSS Selector به‌روزرسانی شده
    try:
        price_element = driver.find_element(By.CSS_SELECTOR, "#app > div.container--has-footer-d86a9.kt-container > div > main > article > div > div.kt-col-5 > section:nth-child(1) > div.post-page__section--padded > div:nth-child(13)")
        ad_data['قیمت آگهی (تومان)'] = price_element.text.strip()
        print(f"✅ قیمت آگهی از CSS Selector به‌روزرسانی شده: {ad_data['قیمت آگهی (تومان)']}")
    except Exception as e:
        print(f"⚠️ خطا در استخراج قیمت آگهی از CSS Selector به‌روزرسانی شده: {e}")
        # fallback به سلکتور قبلی
        try:
            price_element_old = driver.find_element(By.CSS_SELECTOR, "#app > div.container--has-footer-d86a9.kt-container > div > main > article > div > div.kt-col-5 > section:nth-child(1) > div.post-page__section--padded > div:nth-child(11) > div.kt-base-row__end.kt-unexpandable-row__value-box")
            ad_data['قیمت آگهی (تومان)'] = price_element_old.text.strip()
            print(f"✅ قیمت آگهی از CSS Selector قبلی: {ad_data['قیمت آگهی (تومان)']}")
        except Exception as e2:
            print(f"⚠️ خطا در استخراج قیمت آگهی از CSS Selector قبلی: {e2}")
            # fallback به روش‌های مختلف
            try:
                # تلاش با سلکتورهای مختلف قیمت
                price_selectors = [
                    ".kt-unexpandable-row__value",
                    ".kt-base-row__end",
                    "[class*='price']",
                    "[class*='قیمت']"
                ]
                
                for selector in price_selectors:
                    try:
                        price_elements = driver.find_elements(By.CSS_SELECTOR, selector)
                        for elem in price_elements:
                            text = elem.text.strip()
                            if text and ('تومان' in text or 'میلیون' in text or text.replace(',', '').isdigit()):
                                ad_data['قیمت آگهی (تومان)'] = text
                                print(f"✅ قیمت از fallback selector: {text}")
                                break
                        if ad_data.get('قیمت آگهی (تومان)'):
                            break
                    except:
                        continue
                        
                # اگر هنوز قیمت پیدا نشد، از متن صفحه استخراج کن
                if not ad_data.get('قیمت آگهی (تومان)'):
                    page_text = driver.find_element(By.TAG_NAME, 'body').text
                    price_patterns = [
                        r'(\d{1,3}(?:,\d{3})*\s*(?:میلیون|تومان))',
                        r'قیمت[:\s]*(\d{1,3}(?:,\d{3})*)',
                        r'(\d{1,3}(?:,\d{3})*(?:\s*میلیون)?\s*تومان)'
                    ]
                    for pattern in price_patterns:
                        matches = re.findall(pattern, page_text)
                        if matches:
                            ad_data['قیمت آگهی (تومان)'] = matches[0]
                            print(f"✅ قیمت از متن صفحه: {matches[0]}")
                            break
            except:
                ad_data['قیمت آگهی (تومان)'] = 'استخراج نشد'
    
    # استخراج سریع اطلاعات با محدودیت تعداد ردیف‌ها
    try:
        # یک بار تمام ردیف‌ها را پیدا کن - محدود به 10 ردیف اول برای سرعت
        rows = driver.find_elements(By.CSS_SELECTOR, ".kt-base-row, .kt-unexpandable-row")[:10]
        print(f"✅ بررسی {len(rows)} ردیف اول برای سرعت بیشتر")
        
        # دیکشنری برای نگهداری کلمات کلیدی
        field_keywords = {
            'وضعیت موتور': ['وضعیت موتور', 'موتور'],
            'وضعیت شاسی': ['وضعیت شاسی', 'شاسی'],
            'وضعیت بدنه': ['وضعیت بدنه', 'بدنه'],
            'گیربکس': ['گیربکس', 'دنده'],
            'قیمت آگهی (تومان)': ['قیمت'],
            'بیمه شخص ثالث': ['بیمه'],
            'معاوضه': ['معاوضه']
        }
        
        # پردازش سریع ردیف‌ها
        for row in rows:
            try:
                row_text = row.text.strip()
                if not row_text or len(row_text) < 3:
                    continue
                
                # بررسی سریع وجود کلمات کلیدی
                row_lower = row_text.lower()
                matched_field = None
                
                for field, keywords in field_keywords.items():
                    if any(keyword in row_lower for keyword in keywords):
                        matched_field = field
                        break
                
                if matched_field and not ad_data.get(matched_field):
                    # استخراج مقدار با روش بهینه
                    if ':' in row_text:
                        parts = row_text.split(':', 1)
                        if len(parts) == 2:
                            value = parts[1].strip()
                            if value and len(value) > 1:
                                ad_data[matched_field] = value
                                print(f"✅ {matched_field}: {value}")
                    else:
                        # تلاش برای استخراج با selector
                        try:
                            value_elem = row.find_element(By.CSS_SELECTOR, ".kt-base-row__end, .kt-unexpandable-row__value")
                            value = value_elem.text.strip()
                            if value and len(value) > 1:
                                ad_data[matched_field] = value
                                print(f"✅ {matched_field}: {value}")
                        except:
                            continue
                            
            except Exception as e:
                continue
                
    except Exception as e:
        print(f"⚠️ خطا در استخراج اطلاعات: {e}")
    
    # استخراج سریع وضعیت موتور - فقط سلکتورهای اصلی
    if not ad_data.get('وضعیت موتور'):
        engine_selectors = [
            "#app > div.container--has-footer-d86a9.kt-container > div > main > article > div > div.kt-col-5 > section:nth-child(1) > div.post-page__section--padded > div:nth-child(12)",
            ".kt-base-row__end"
        ]
        
        for selector in engine_selectors[:2]:  # فقط 2 سلکتور اول
            try:
                elements = driver.find_elements(By.CSS_SELECTOR, selector)[:3]  # فقط 3 المنت اول
                for element in elements:
                    text = element.text.strip()
                    if text and 'موتور' not in text and len(text) > 2:
                        ad_data['وضعیت موتور'] = text
                        break
                if ad_data.get('وضعیت موتور'):
                    break
            except:
                continue
    
    # بهبود استخراج وضعیت شاسی با روش‌های متنوع
    if not ad_data.get('وضعیت شاسی'):
        chassis_selectors = [
            "div[class*='chassis'], div[class*='frame']",
            "*[class*='unexpandable-row']:has(*:contains('شاسی'))",
            "div:contains('وضعیت شاسی') + div",
            "div:contains('شاسی') ~ div",
            ".kt-base-row:has(*:contains('شاسی')) .kt-base-row__end",
            "#app > div.container--has-footer-d86a9.kt-container > div > main > article > div > div.kt-col-5 > section:nth-child(1) > div.post-page__section--padded > div:nth-child(14)"
        ]
        
        for selector in chassis_selectors:
            try:
                elements = driver.find_elements(By.CSS_SELECTOR, selector)
                for element in elements:
                    text = element.text.strip()
                    if text and 'شاسی' not in text and len(text) > 2:
                        ad_data['وضعیت شاسی'] = text
                        print(f"✅ وضعیت شاسی از selector {selector}: {text}")
                        break
                if ad_data.get('وضعیت شاسی'):
                    break
            except:
                continue
    
    # بهبود استخراج وضعیت بدنه با روش‌های متنوع
    if not ad_data.get('وضعیت بدنه'):
        body_selectors = [
            "div[class*='body'], div[class*='paint']",
            "*[class*='unexpandable-row']:has(*:contains('بدنه'))",
            "div:contains('وضعیت بدنه') + div",
            "div:contains('بدنه') ~ div",
            ".kt-base-row:has(*:contains('بدنه')) .kt-base-row__end",
            "#app > div.container--has-footer-d86a9.kt-container > div > main > article > div > div.kt-col-5 > section:nth-child(1) > div.post-page__section--padded > div:nth-child(16)"
        ]
        
        for selector in body_selectors:
            try:
                elements = driver.find_elements(By.CSS_SELECTOR, selector)
                for element in elements:
                    text = element.text.strip()
                    if text and 'بدنه' not in text and len(text) > 2:
                        ad_data['وضعیت بدنه'] = text
                        print(f"✅ وضعیت بدنه از selector {selector}: {text}")
                        break
                if ad_data.get('وضعیت بدنه'):
                    break
            except:
                continue
    
    # روش قدیمی برای گیربکس اگر با روش جدید پیدا نشد
    if 'گیربکس' not in ad_data:
        try:
            gearbox_element = driver.find_element(By.CSS_SELECTOR, "#app > div.container--has-footer-d86a9.kt-container > div > main > article > div > div.kt-col-5 > section:nth-child(1) > div.post-page__section--padded > div:nth-child(5)")
            ad_data['گیربکس'] = gearbox_element.text.strip()
            print(f"✅ گیربکس از CSS Selector جدید: {ad_data['گیربکس']}")
        except Exception as e:
            print(f"⚠️ خطا در استخراج گیربکس از CSS Selector جدید: {e}")
    
    # برند و تیپ - روش قبلی
    try:
        rows = driver.find_elements(By.CSS_SELECTOR, ".kt-unexpandable-row")
        for row in rows[:5]:  # فقط 5 ردیف اول برای سرعت
            try:
                title_elem = row.find_element(By.CSS_SELECTOR, ".kt-unexpandable-row__title")
                title = title_elem.text.strip()
                
                if 'برند و تیپ' in title:
                    value_elem = row.find_element(By.CSS_SELECTOR, ".kt-unexpandable-row__action")
                    ad_data['برند و تیپ'] = value_elem.text.strip()
            except:
                continue
    except:
        pass
    
    # استخراج بهینه‌شده توضیحات
    if not ad_data.get('توضیحات'):
        try:
            # selector های اصلی و مؤثر
            primary_selectors = [
                ".kt-description-row", 
                "[class*='description']", 
                "div.post-page__section--padded",
                "textarea"
            ]
            
            # کلمات کلیدی برای تشخیص توضیحات معتبر
            valid_keywords = ['سالم', 'تعمیر', 'رنگ', 'موتور', 'گیربکس', 'خودرو', 'ماشین']
            
            for selector in primary_selectors:
                try:
                    elements = driver.find_elements(By.CSS_SELECTOR, selector)
                    for element in elements:
                        desc = element.text.strip()
                        
                        # بررسی سریع کیفیت
                        if (desc and len(desc) > 20 and 
                            (len(desc) > 30 or any(keyword in desc.lower() for keyword in valid_keywords))):
                            
                            # بررسی عدم تکرار
                            title = ad_data.get('عنوان آگهی', '')
                            if desc != title and desc != 'بدون توضیحات':
                                ad_data['توضیحات'] = desc
                                print(f"✅ توضیحات: {desc[:40]}...")
                                break
                    
                    if ad_data.get('توضیحات'):
                        break
                except:
                    continue
                    
        except Exception as e:
            print(f"⚠️ خطا در استخراج توضیحات: {e}")
    
    # روش 2: جستجو برای بخش توضیحات با عنوان در تمام المان‌های صفحه
    if 'توضیحات' not in ad_data:
        try:
            # جستجو برای تمام المان‌های ممکن که می‌توانند عنوان بخش باشند
            title_elements = driver.find_elements(By.XPATH, "//*[contains(text(), 'توضیحات')]")
            
            for title_elem in title_elements:
                try:
                    # بررسی المان والد یا بعدی برای یافتن متن توضیحات
                    parent = title_elem.find_element(By.XPATH, "..")
                    desc = parent.text.replace(title_elem.text, "").strip()
                    
                    if desc and len(desc) > 10:
                        ad_data['توضیحات'] = desc
                        print(f"✅ توضیحات از روش جستجوی متنی: {desc[:30]}...")
                        break
                        
                    # بررسی المان‌های بعدی
                    next_elements = parent.find_elements(By.XPATH, "./following-sibling::*")
                    for next_elem in next_elements[:3]:  # بررسی حداکثر 3 المان بعدی
                        desc = next_elem.text.strip()
                        if desc and len(desc) > 10:
                            ad_data['توضیحات'] = desc
                            print(f"✅ توضیحات از المان بعدی: {desc[:30]}...")
                            break
                    
                    if 'توضیحات' in ad_data:
                        break
                except:
                    continue
        except Exception as e:
            print(f"⚠️ خطا در استخراج توضیحات از روش جستجوی متنی: {e}")
    
    # روش 3: جستجو در تمام بخش‌های صفحه
    if 'توضیحات' not in ad_data:
        try:
            sections = driver.find_elements(By.CSS_SELECTOR, "section, div.post-page__section--padded, div.kt-base-row")
            for section in sections:
                try:
                    section_text = section.text.strip()
                    if 'توضیحات' in section_text and len(section_text) > 20:
                        # حذف عنوان از متن بخش
                        lines = section_text.split('\n')
                        if len(lines) > 1:
                            desc = '\n'.join(lines[1:]).strip()
                            if desc:
                                ad_data['توضیحات'] = desc
                                print(f"✅ توضیحات از بخش صفحه: {desc[:30]}...")
                                break
                except:
                    continue
        except Exception as e:
            print(f"⚠️ خطا در استخراج توضیحات از بخش‌های صفحه: {e}")
    
    # اگر هیچ توضیحاتی پیدا نشد، مقدار پیش‌فرض قرار می‌دهیم
    if 'توضیحات' not in ad_data:
        ad_data['توضیحات'] = "بدون توضیحات"
    
    # زمان و مکان با روش عمومی و جامع
    if 'زمان و مکان' not in ad_data:
        try:
            # روش 1: جستجو برای تمام المان‌های ممکن زمان و مکان با کلاس‌های مختلف
            location_selectors = [
                ".kt-page-title__subtitle", 
                ".kt-page-title div", 
                "[class*='subtitle']", 
                "[class*='location']", 
                "[class*='time']",
                "div.kt-page-title",
                "div.post-header__subtitle"
            ]
            
            for selector in location_selectors:
                try:
                    location_elements = driver.find_elements(By.CSS_SELECTOR, selector)
                    for element in location_elements:
                        location_text = element.text.strip()
                        # بررسی اینکه متن شامل کلمات کلیدی زمان و مکان باشد
                        if location_text and ('در' in location_text or 'پیش' in location_text or 'لحظاتی' in location_text or 'ساعت' in location_text or 'دقیقه' in location_text):
                            ad_data['زمان و مکان'] = location_text
                            print(f"✅ زمان و مکان از سلکتور {selector}: {location_text}")
                            break
                    
                    if 'زمان و مکان' in ad_data:
                        break
                except Exception as e:
                    print(f"⚠️ خطا در استخراج زمان و مکان از سلکتور {selector}: {e}")
                    continue
        except Exception as e:
            print(f"⚠️ خطا در استخراج زمان و مکان از روش عمومی: {e}")
    
    # روش 2: جستجو برای متن حاوی زمان و مکان در تمام المان‌های صفحه
    if 'زمان و مکان' not in ad_data:
        try:
            # جستجو برای تمام المان‌های متنی در صفحه
            text_elements = driver.find_elements(By.XPATH, "//div[contains(text(), 'در') or contains(text(), 'پیش') or contains(text(), 'لحظاتی') or contains(text(), 'ساعت') or contains(text(), 'دقیقه')]")
            
            for element in text_elements:
                try:
                    location_text = element.text.strip()
                    if location_text and len(location_text) < 100:  # اطمینان از اینکه متن خیلی طولانی نباشد
                        ad_data['زمان و مکان'] = location_text
                        print(f"✅ زمان و مکان از جستجوی متنی: {location_text}")
                        break
                except:
                    continue
        except Exception as e:
            print(f"⚠️ خطا در استخراج زمان و مکان از جستجوی متنی: {e}")
    
    # روش 3: جستجو در هدر صفحه
    if 'زمان و مکان' not in ad_data:
        try:
            header_elements = driver.find_elements(By.CSS_SELECTOR, "header *, .kt-page-title *, .post-header *")
            for element in header_elements:
                try:
                    location_text = element.text.strip()
                    if location_text and ('در' in location_text or 'پیش' in location_text or 'لحظاتی' in location_text):
                        ad_data['زمان و مکان'] = location_text
                        print(f"✅ زمان و مکان از هدر صفحه: {location_text}")
                        break
                except:
                    continue
        except Exception as e:
            print(f"⚠️ خطا در استخراج زمان و مکان از هدر صفحه: {e}")
    
    # روش 4: CSS Selector دقیق برای زمان و مکان
    if 'زمان و مکان' not in ad_data:
        try:
            location_element = driver.find_element(By.CSS_SELECTOR, "#app > div.container--has-footer-d86a9.kt-container > div > main > article > div > div.kt-col-5 > section:nth-child(1) > div.kt-page-title > div > div")
            ad_data['زمان و مکان'] = location_element.text.strip()
            print(f"✅ زمان و مکان از CSS Selector دقیق: {ad_data['زمان و مکان']}")
        except Exception as e:
            print(f"⚠️ خطا در استخراج زمان و مکان از CSS Selector دقیق: {e}")
    
    # اگر هیچ زمان و مکانی پیدا نشد، مقدار پیش‌فرض قرار می‌دهیم
    if 'زمان و مکان' not in ad_data:
        ad_data['زمان و مکان'] = "نامشخص"
    
    # استخراج دقیق نام خودرو و برند و تیپ
    title = ad_data.get('عنوان آگهی', '')
    brand_type = ad_data.get('برند و تیپ', '')
    
    # دیکشنری برندهای خودرو با کلمات کلیدی
    car_brands = {
        'پراید': ['پراید', 'pride'],
        'پژو': ['پژو', 'peugeot', '206', '207', '405', '406', 'پارس', 'پرشیا'],
        'سمند': ['سمند', 'samand'],
        'دنا': ['دنا', 'dena'],
        'رانا': ['رانا', 'rana'],
        'تیبا': ['تیبا', 'tiba'],
        'ساینا': ['ساینا', 'saina'],
        'کوییک': ['کوییک', 'quick'],
        'آریو': ['آریو', 'ario'],
        'شاهین': ['شاهین', 'shahin'],
        'تارا': ['تارا', 'tara'],
        'تویوتا': ['تویوتا', 'toyota', 'کمری', 'کرولا', 'پرادو'],
        'هوندا': ['هوندا', 'honda', 'سیویک', 'آکورد'],
        'نیسان': ['نیسان', 'nissan'],
        'هیوندای': ['هیوندای', 'hyundai', 'النترا', 'سوناتا'],
        'کیا': ['کیا', 'kia', 'سراتو', 'اپتیما'],
        'مزدا': ['مزدا', 'mazda'],
        'فولکس': ['فولکس', 'volkswagen', 'پاسات', 'جتا'],
        'رنو': ['رنو', 'renault', 'ساندرو', 'تندر'],
        'چری': ['چری', 'chery', 'تیگو'],
        'ام وی ام': ['ام وی ام', 'mvm', '110', '315', '530'],
        'بنز': ['بنز', 'mercedes', 'مرسدس'],
        'بی ام و': ['بی ام و', 'bmw'],
        'سوزوکی': ['سوزوکی', 'suzuki', 'ویتارا', 'سوئیفت'],
        'جک': ['جک', 'jac'],
        'لیفان': ['لیفان', 'lifan'],
        'دانگ فنگ': ['دانگ فنگ', 'dongfeng'],
        'جیلی': ['جیلی', 'geely']
    }
    
    # تشخیص برند از عنوان
    detected_brand = ''
    for brand, keywords in car_brands.items():
        for keyword in keywords:
            if keyword.lower() in title.lower():
                detected_brand = brand
                break
        if detected_brand:
            break
    
    # اگر برند از brand_type استخراج شده، آن را استفاده کن
    if brand_type and brand_type != title:
        ad_data['برند و تیپ'] = brand_type
        # تشخیص برند از brand_type
        if not detected_brand:
            for brand, keywords in car_brands.items():
                for keyword in keywords:
                    if keyword.lower() in brand_type.lower():
                        detected_brand = brand
                        break
                if detected_brand:
                    break
    else:
        # استخراج برند و تیپ از عنوان
        if detected_brand:
            # استخراج مدل و سال از عنوان
            title_parts = title.split()
            model_parts = []
            for part in title_parts:
                if not any(keyword.lower() in part.lower() for keywords in car_brands.values() for keyword in keywords):
                    if part.isdigit() and len(part) == 4 and 1300 <= int(part) <= 1410:  # سال شمسی
                        continue
                    if 'مدل' not in part.lower():
                        model_parts.append(part)
            
            if model_parts:
                ad_data['برند و تیپ'] = f"{detected_brand} {' '.join(model_parts[:2])}"
            else:
                ad_data['برند و تیپ'] = detected_brand
        else:
            ad_data['برند و تیپ'] = title.split()[0] if title.split() else ''
    
    # تنظیم نام خودرو
    ad_data['نام خودرو'] = detected_brand if detected_brand else (title.split()[0] if title.split() else '')
    
    return ad_data


def check_engine_gearbox_issues(description):
    """تشخیص مشکلات موتور و گیربکس از متن توضیحات با کلمات کلیدی گسترده‌تر"""
    if not description:
        return False, []
    
    description_lower = description.lower()
    
    # کلمات کلیدی مشکلات موتور (گسترده‌تر)
    engine_issues = [
        'موتور تعمیر', 'موتور تعویض', 'موتور خراب', 'موتور معیوب',
        'تعمیر موتور', 'تعویض موتور', 'خرابی موتور', 'مشکل موتور',
        'موتور آسیب', 'موتور ضربه', 'موتور سوخته', 'اورهال موتور',
        'موتور اورهال', 'بازسازی موتور', 'موتور بازسازی', 'موتور تازه تعمیر',
        'موتور نو', 'موتور جدید', 'تعمیر کامل موتور', 'موتور کامل تعمیر',
        'موتور نیاز به تعمیر', 'موتور مشکل دار', 'موتور دود می‌کند',
        'موتور صدا دارد', 'موتور لرزش', 'موتور ویبره', 'موتور گرم می‌کند'
    ]
    
    # کلمات کلیدی مشکلات گیربکس (گسترده‌تر)
    gearbox_issues = [
        'گیربکس تعمیر', 'گیربکس تعویض', 'گیربکس خراب', 'گیربکس معیوب',
        'تعمیر گیربکس', 'تعویض گیربکس', 'خرابی گیربکس', 'مشکل گیربکس',
        'گیربکس آسیب', 'گیربکس ضربه', 'گیربکس سوخته', 'اورهال گیربکس',
        'گیربکس اورهال', 'بازسازی گیربکس', 'گیربکس بازسازی', 'گیربکس تازه تعمیر',
        'گیربکس نو', 'گیربکس جدید', 'تعمیر کامل گیربکس', 'گیربکس کامل تعمیر',
        'گیربکس نیاز به تعمیر', 'گیربکس مشکل دار', 'گیربکس صدا دارد',
        'گیربکس سخت می‌گیرد', 'دنده سخت', 'دنده نمی‌گیرد', 'کلاچ تعمیر',
        'کلاچ خراب', 'کلاچ معیوب', 'مشکل کلاچ', 'کلاچ سوخته'
    ]
    
    found_issues = []
    
    # بررسی مشکلات موتور
    for issue in engine_issues:
        if issue in description_lower:
            found_issues.append(f"موتور: {issue}")
    
    # بررسی مشکلات گیربکس
    for issue in gearbox_issues:
        if issue in description_lower:
            found_issues.append(f"گیربکس: {issue}")
    
    has_issues = len(found_issues) > 0
    return has_issues, found_issues


def check_urgent_sale_keywords(description):
    """تشخیص کلمات فوری فروش از متن توضیحات"""
    if not description:
        return False
    
    description_lower = description.lower()
    
    # کلمات کلیدی فروش فوری
    urgent_keywords = [
        'پول لازم', 'فروش فوری', 'نیاز مالی', 'زیر قیمت',
        'فوری فروش', 'نقدی فوری', 'ضروری فروش', 'سریع فروش',
        'گیر پولم','عجله دارم', 'فوری نقد', 'قیمت پایین', 'ارزان فروش'
    ]
    
    for keyword in urgent_keywords:
        if keyword in description_lower:
            return True
    
    return False


# کد همراه مکانیک و Z4Car حذف شد و با combined_scraper.py جایگزین شده است
# توابع scrape_hamrah_mechanic و scrape_z4car از combined_scraper import می‌شوند


def process_ad_with_pricing(ad_data):
    """پردازش آگهی با قیمت‌گذاری پیشرفته و محاسبه افت قیمت"""
    try:
        print(f"💰 پردازش آگهی: {ad_data['عنوان آگهی'][:50]}...")
        
        # بررسی مشکلات موتور و گیربکس
        description = ad_data.get('توضیحات', '')
        has_engine_issues, engine_issues_list = check_engine_gearbox_issues(description)
        
        # بررسی کلمات فوری فروش
        is_urgent_sale = check_urgent_sale_keywords(description)
        
        # اضافه کردن فیلدهای جدید
        ad_data['مشکلات موتور/گیربکس'] = ', '.join(engine_issues_list) if engine_issues_list else 'سالم'
        ad_data['فروش فوری'] = 'بله' if is_urgent_sale else 'خیر'
        ad_data['رنگ‌بندی'] = 'سبز' if is_urgent_sale else 'عادی'
        
        # استخراج نام خودرو از برند و تیپ
        brand_type = ad_data.get('برند و تیپ', '')
        if brand_type:
            # تشخیص نام خودرو از برند و تیپ
            if 'پراید' in brand_type:
                ad_data['نام خودرو'] = 'پراید'
            elif 'پژو' in brand_type:
                ad_data['نام خودرو'] = 'پژو'
            elif 'سمند' in brand_type:
                ad_data['نام خودرو'] = 'سمند'
            elif 'دنا' in brand_type:
                ad_data['نام خودرو'] = 'دنا'
            elif 'رانا' in brand_type:
                ad_data['نام خودرو'] = 'رانا'
            elif 'تیبا' in brand_type:
                ad_data['نام خودرو'] = 'تیبا'
            else:
                # استخراج اولین کلمه به عنوان نام خودرو
                ad_data['نام خودرو'] = brand_type.split()[0] if brand_type.split() else 'نامشخص'
        
        # محاسبه قیمت روز با دقت بیشتر از برند + تیپ + سال
        market_price = None
        if IMPROVED_PRICING_AVAILABLE:
            try:
                calculator = ImprovedCarPriceCalculator()
                
                # بارگذاری قیمت‌های بازار
                try:
                    market_df = pd.read_excel('combined_market_prices.xlsx')
                    calculator.load_market_prices('combined_market_prices.xlsx')
                except:
                    market_df = None
                
                # استخراج اطلاعات خودرو برای جستجوی دقیق‌تر
                title = ad_data.get('عنوان آگهی', '')
                description = ad_data.get('توضیحات', '')
                brand_type = ad_data.get('برند و تیپ', '')
                
                car_info = calculator.extract_car_info(title, description)
                
                # جستجوی قیمت با اطلاعات کامل
                market_price = calculator.find_market_price(car_info, market_df, brand_type)
                
                if market_price:
                    ad_data['قیمت روز (تومان)'] = f"{market_price:,} تومان"
                    ad_data['منبع قیمت'] = 'همراه مکانیک + زد فور کار (دقیق)'
                else:
                    # fallback به روش قبلی
                    car_name = brand_type or ad_data.get('نام خودرو') or title
                    if car_name:
                        market_price_data = get_market_price_for_car(car_name, force_update=False)
                        market_price = market_price_data.get('market_price')
                        if market_price:
                            ad_data['قیمت روز (تومان)'] = f"{market_price:,} تومان"
                            ad_data['منبع قیمت'] = 'همراه مکانیک + زد فور کار'
                        else:
                            ad_data['قیمت روز (تومان)'] = 'یافت نشد'
                            ad_data['منبع قیمت'] = 'دیوار'
                    else:
                        ad_data['قیمت روز (تومان)'] = 'نامشخص'
                        ad_data['منبع قیمت'] = 'دیوار'
                        
            except Exception as e:
                print(f"⚠️ خطا در جستجوی دقیق قیمت: {e}")
                # fallback به روش قبلی
                car_name = ad_data.get('برند و تیپ') or ad_data.get('نام خودرو') or ad_data.get('عنوان آگهی')
                if car_name:
                    market_price_data = get_market_price_for_car(car_name, force_update=False)
                    market_price = market_price_data.get('market_price')
                    if market_price:
                        ad_data['قیمت روز (تومان)'] = f"{market_price:,} تومان"
                        ad_data['منبع قیمت'] = 'همراه مکانیک + زد فور کار'
                    else:
                        ad_data['قیمت روز (تومان)'] = 'یافت نشد'
                        ad_data['منبع قیمت'] = 'دیوار'
                else:
                    ad_data['قیمت روز (تومان)'] = 'نامشخص'
                    ad_data['منبع قیمت'] = 'دیوار'
        else:
            # روش قبلی اگر سیستم بهبود یافته در دسترس نباشد
            car_name = ad_data.get('برند و تیپ') or ad_data.get('نام خودرو') or ad_data.get('عنوان آگهی')
            if car_name:
                market_price_data = get_market_price_for_car(car_name, force_update=False)
                market_price = market_price_data.get('market_price')
                if market_price:
                    ad_data['قیمت روز (تومان)'] = f"{market_price:,} تومان"
                    ad_data['منبع قیمت'] = 'همراه مکانیک + زد فور کار'
                else:
                    ad_data['قیمت روز (تومان)'] = 'یافت نشد'
                    ad_data['منبع قیمت'] = 'دیوار'
            else:
                ad_data['قیمت روز (تومان)'] = 'نامشخص'
                ad_data['منبع قیمت'] = 'دیوار'
        
        # محاسبه درصدهای افت قیمت با ImprovedCarPriceCalculator
        if IMPROVED_PRICING_AVAILABLE:
            try:
                calculator = ImprovedCarPriceCalculator()
                
                # استخراج اطلاعات خودرو
                title = ad_data.get('عنوان آگهی', '')
                car_info = calculator.extract_car_info(title, description)
                
                # تشخیص مشکلات با در نظر گیری وضعیت‌های مختلف
                body_status = ad_data.get('وضعیت بدنه', '')
                engine_status = ad_data.get('وضعیت موتور', '')
                gearbox_status = ad_data.get('وضعیت گیربکس', '')
                chassis_status = ad_data.get('وضعیت شاسی', '')
                
                issues = calculator.detect_issues(
                    title, description, body_status, engine_status, gearbox_status, chassis_status
                )
                
                # استخراج کیلومتر و سال از ad_data یا car_info
                mileage = ad_data.get('کیلومتر') or car_info.get('mileage')
                year = ad_data.get('سال') or car_info.get('year')
                
                # تبدیل رشته کیلومتر به عدد
                if isinstance(mileage, str):
                    mileage_num = re.sub(r'[^\d]', '', mileage)
                    mileage = int(mileage_num) if mileage_num else None
                
                # تبدیل رشته سال به عدد
                if isinstance(year, str):
                    year_num = re.sub(r'[^\d]', '', year)
                    year = int(year_num) if year_num else None
                
                # محاسبه افت کارکرد
                mileage_depreciation = calculator.calculate_mileage_depreciation(mileage, year)
                
                # محاسبه افت سن خودرو
                age_depreciation = calculator.calculate_age_depreciation(year)
                
                # محاسبه کل افت قیمت
                total_depreciation, issues_depreciation = calculator.calculate_total_depreciation(
                    issues, mileage_depreciation, age_depreciation
                )
                
                # محاسبه قیمت تخمینی
                estimated_price = calculator.calculate_estimated_price(market_price, total_depreciation)
                
                # به‌روزرسانی داده‌ها
                ad_data['مشکلات تشخیص داده شده'] = ', '.join(issues) if issues else 'هیچ'
                ad_data['افت کارکرد'] = f"{mileage_depreciation:.1%}"
                ad_data['افت سن خودرو'] = f"{age_depreciation:.1%}"
                ad_data['افت مشکلات'] = f"{issues_depreciation:.1%}"
                ad_data['درصد افت کل'] = f"{total_depreciation:.1%}"
                ad_data['قیمت تخمینی (تومان)'] = f"{estimated_price:,.0f} تومان" if estimated_price else 'محاسبه نشد'
                
                print(f"📊 کیلومتر: {mileage:,}, سال: {year}, افت کارکرد: {mileage_depreciation:.1%}, افت سن: {age_depreciation:.1%}, افت کل: {total_depreciation:.1%}" if mileage and year else "📊 اطلاعات ناکافی برای محاسبه")
                print(f"💰 قیمت تخمینی: {estimated_price:,.0f} تومان" if estimated_price else "💰 محاسبه قیمت ناموفق")
                
            except Exception as e:
                print(f"⚠️ خطا در محاسبه افت قیمت: {e}")
                # مقادیر پیش‌فرض در صورت خطا
                ad_data['مشکلات تشخیص داده شده'] = ', '.join(engine_issues_list) if engine_issues_list else 'هیچ'
                ad_data['افت کارکرد'] = '0.0%'
                ad_data['افت سن خودرو'] = '0.0%'
                ad_data['افت مشکلات'] = '0.0%'
                ad_data['درصد افت کل'] = '0.0%'
                ad_data['قیمت تخمینی (تومان)'] = 'خطا در محاسبه'
        else:
            # مقادیر پیش‌فرض اگر سیستم بهبود یافته در دسترس نباشد
            ad_data['مشکلات تشخیص داده شده'] = ', '.join(engine_issues_list) if engine_issues_list else 'هیچ'
            ad_data['افت کارکرد'] = '0.0%'
            ad_data['افت سن خودرو'] = '0.0%'
            ad_data['افت مشکلات'] = '0.0%'
            ad_data['درصد افت کل'] = '0.0%'
            ad_data['قیمت تخمینی (تومان)'] = 'سیستم محاسبه در دسترس نیست'
        
        # اضافه کردن تاریخ ذخیره
        ad_data['تاریخ ذخیره'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        print(f"✅ پردازش کامل شد")
        return ad_data
        
    except Exception as e:
        print(f"❌ خطا در پردازش: {e}")
        return ad_data


def getads():
    """دریافت لیست آگهی‌ها با اسکرول برای آگهی‌های بیشتر"""
    ads = []
    unique_ads = set()
    
    try:
        # انتظار برای بارگذاری اولیه - کاهش timeout برای سرعت بیشتر
        WebDriverWait(driver, 3).until(
            EC.presence_of_element_located((By.TAG_NAME, "article"))
        )
        
        print("🔄 شروع اسکرول برای یافتن آگهی‌های بیشتر...")
        
        # اسکرول برای بارگذاری آگهی‌های بیشتر
        last_count = 0
        scroll_attempts = 0
        max_scrolls = 5  # حداکثر 5 بار اسکرول
        
        while scroll_attempts < max_scrolls:
            # جمع‌آوری آگهی‌های فعلی
            current_ads = driver.find_elements(By.TAG_NAME, "article")
            
            for ad_element in current_ads:
                try:
                    link_element = ad_element.find_element(By.TAG_NAME, "a")
                    href = link_element.get_attribute('href')
                    if href and '/v/' in href and href not in unique_ads:
                        unique_ads.add(href)
                except:
                    continue
            
            current_count = len(unique_ads)
            print(f"📊 آگهی‌های یافت شده: {current_count}")
            
            # اگر آگهی جدیدی پیدا نشد، توقف
            if current_count == last_count:
                scroll_attempts += 1
                if scroll_attempts >= 2:  # اگر 2 بار پشت سر هم آگهی جدید پیدا نشد
                    break
            else:
                scroll_attempts = 0  # reset counter
            
            last_count = current_count
            
            # اسکرول به پایین
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(1.5)  # کاهش زمان انتظار برای سرعت بیشتر
            
            # بررسی وجود دکمه "آگهی بیشتر" و کلیک روی آن
            try:
                more_ads_button = driver.find_element(By.CSS_SELECTOR, "#post-list-container-id > div.post-list__bottom-container-cac2f > div")
                if more_ads_button.is_displayed():
                    driver.execute_script("arguments[0].click();", more_ads_button)
                    print("🔄 کلیک روی دکمه آگهی بیشتر")
                    time.sleep(2)  # انتظار برای بارگذاری آگهی‌های جدید
            except:
                pass  # اگر دکمه وجود نداشت، ادامه بده
            
            # اگر به تعداد مطلوب رسیدیم، توقف
            if current_count >= 300:  # حداکثر 300 آگهی
                break
        
        ad_links = list(unique_ads)
        print(f"✅ مجموع {len(ad_links)} آگهی منحصر به فرد یافت شد")
        
        return ad_links
        
    except Exception as e:
        print(f"❌ خطا در دریافت آگهی‌ها: {e}")
        # fallback به روش ساده
        try:
            ad_elements = driver.find_elements(By.TAG_NAME, "article")
            ad_links = []
            for ad_element in ad_elements:
                try:
                    link_element = ad_element.find_element(By.TAG_NAME, "a")
                    href = link_element.get_attribute('href')
                    if href and '/v/' in href:
                        ad_links.append(href)
                except:
                    continue
            return ad_links
        except:
            return []


# تابع cleanup برای ذخیره ایمن داده‌ها
def cleanup_and_exit(reason="نامشخص"):
    """ذخیره ایمن داده‌های باقی‌مانده و خروج از برنامه"""
    global processed_count, data_batch, record_count
    
    print(f"\n🛑 دریافت سیگنال توقف ({reason})...")
    print("💾 در حال ذخیره داده‌های باقی‌مانده...")
    
    # ذخیره batch باقی‌مانده
    if data_batch:
        try:
            save_batch_to_improved_file()
            print(f"💾 ذخیره نهایی: {len(data_batch)} رکورد باقی‌مانده ذخیره شد")
        except Exception as e:
            print(f"❌ خطا در ذخیره نهایی: {e}")
            # تلاش برای ذخیره اضطراری
            try:
                from datetime import datetime
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                emergency_file = f"divar_ads_emergency_{timestamp}.xlsx"
                df = pd.DataFrame(data_batch)
                df.to_excel(emergency_file, index=False, engine='openpyxl')
                print(f"🚨 ذخیره اضطراری در {emergency_file}")
            except:
                print("❌ خطا در ذخیره اضطراری")
    else:
        print("⚠️ هیچ داده باقی‌مانده‌ای برای ذخیره وجود ندارد")
    
    print(f"✅ پردازش کامل شد")
    print(f"📊 مجموع آگهی‌های پردازش شده: {record_count}")
    
    # بستن driver
    try:
        driver.quit()
        print("🔒 Chrome driver بسته شد")
    except:
        pass
    
    print("👋 خروج از برنامه...")
    sys.exit(0)

# تعریف signal handler برای Ctrl+C و سایر سیگنال‌ها
def signal_handler(sig, frame):
    signal_names = {
        signal.SIGINT: "Ctrl+C",
        signal.SIGTERM: "SIGTERM",
        signal.SIGHUP: "SIGHUP" if hasattr(signal, 'SIGHUP') else "نامشخص"
    }
    signal_name = signal_names.get(sig, f"Signal {sig}")
    cleanup_and_exit(signal_name)

# ثبت signal handler برای سیگنال‌های مختلف
signal.signal(signal.SIGINT, signal_handler)  # Ctrl+C
signal.signal(signal.SIGTERM, signal_handler)  # Termination signal
if hasattr(signal, 'SIGHUP'):
    signal.signal(signal.SIGHUP, signal_handler)  # Hangup signal (Unix only)

# تابع atexit برای اطمینان از ذخیره داده‌ها در یک فایل واحد
import atexit

def atexit_handler():
    """handler برای خروج برنامه - فقط اگر داده‌ای وجود داشته باشد"""
    global data_batch
    if data_batch:  # فقط اگر داده‌ای وجود داشته باشد
        cleanup_and_exit("خروج برنامه")

atexit.register(atexit_handler)

# سیستم پردازش بهبود یافته آگهی‌ها
print("🎯 سیستم پردازش بهبود یافته آگهی‌ها")
print("💾 ذخیره خودکار هر 50 آگهی")
print("💡 برای توقف و ذخیره داده‌ها، Ctrl+C را فشار دهید")

# استخراج قیمت‌های روز بازار قبل از شروع
print("🔄 در حال استخراج قیمت‌های روز بازار...")
fetch_market_prices(force_update=True)
print("✅ قیمت‌های روز بازار با موفقیت به‌روز شدند")

# بارگذاری داده‌های موجود
load_existing_data()

print("🚀 شروع پردازش آگهی‌ها...")
print("🔄 پردازش بدون محدودیت - تا زمان توقف دستی")

passed = []
processed_count = 0

restart_attempts = 0
max_restart_attempts = 3

# متغیرهای جدید برای شمارش خطاهای pagination
consecutive_pagination_errors = 0
max_consecutive_pagination_errors = 3

while True:
    try:
        # بررسی graceful_exit در ابتدای هر حلقه
        if graceful_exit:
            print("🛑 خروج درخواست شده - متوقف می‌شویم...")
            break
            
        print(f"🔄 شروع حلقه جدید - رکورد فعلی: {record_count}")
        # انتظار خیلی کوتاه قبل از شروع
        time.sleep(random.uniform(0.3, 1))
        
        print("📋 دریافت لیست آگهی‌ها...")
        ad_links = getads()
        print(f"📊 تعداد آگهی‌های یافت شده: {len(ad_links)}")
        
        new_ads_found = False
        restart_attempts = 0  # Reset restart attempts on successful operation
        
        print(f"🔍 شروع پردازش {len(ad_links)} آگهی...")
        for i, ad_href in enumerate(ad_links, 1):
            try:
                # بررسی graceful_exit در هر آگهی
                if graceful_exit:
                    print("🛑 خروج درخواست شده - متوقف می‌شویم...")
                    break
                    
                print(f"📝 پردازش آگهی {i}/{len(ad_links)}: {ad_href[-20:]}...")
                adId = ad_href.split('/')[-1]

                if adId not in passed:
                    print(f"✅ آگهی جدید یافت شد: {adId}")
                    passed.append(adId)
                    
                    # بررسی فقط مشکلات جدی قبل از کلیک
                    critical_issue = check_for_critical_bot_detection()
                    if critical_issue:
                        alert_user_critical(critical_issue)
                        # تلاش برای بازگشت به صفحه اصلی و ادامه
                        if handle_verification_issue():
                            continue
                        else:
                            break
                    
                    # رفتن مستقیم به آگهی با بهینه‌سازی سرعت
                    print(f"🌐 بارگذاری آگهی: {ad_href}")
                    driver.get(ad_href)
                    
                    # حذف انتظار برای سرعت بیشتر
                    # time.sleep(0.1)  # حذف شد
                    
                    # استخراج سریع اطلاعات
                    print("📊 استخراج اطلاعات...")
                    ad_data = extract_ad_details()
                    
                    if ad_data and ad_data.get('عنوان آگهی'):
                        print(f"✅ استخراج شد: {ad_data['عنوان آگهی'][:50]}...")
                        
                        # پردازش آگهی با قیمت‌گذاری هوشمند
                        print("💰 پردازش قیمت‌گذاری...")
                        processed_ad = process_ad_with_pricing(ad_data)
                        
                        # ذخیره سریع
                        save_to_excel(processed_ad)
                        
                        processed_count += 1
                        new_ads_found = True
                        print(f"✅ پیشرفت: {record_count} آگهی پردازش شده")
                    else:
                        print("❌ خطا در استخراج")
                        
                    # برگشت سریع (کاهش انتظار)
                    driver.back()
                    time.sleep(0.1)
                else:
                    print(f"⏭️ آگهی قبلاً پردازش شده: {adId}")
                    continue
                            
                    # ریست شمارنده خطاها
                    consecutive_pagination_errors = 0
                    
            except Exception as e:
                print(f"Error processing ad: {e}")
                # انتظار در صورت خطا
                time.sleep(random.uniform(8, 15))
                continue
        
        # اگر آگهی جدیدی پیدا نشد، اسکرول کن تا آگهی‌های بیشتر بارگذاری شوند
        if not new_ads_found:
            print("🔄 آگهی جدیدی پیدا نشد، اسکرول برای بارگذاری آگهی‌های بیشتر...")
            consecutive_pagination_errors += 1
            
            # بررسی اینکه آیا به حداکثر خطاهای متوالی رسیده‌ایم
            if consecutive_pagination_errors >= max_consecutive_pagination_errors:
                print(f"❌ بعد از {max_consecutive_pagination_errors} خطای متوالی، بات متوقف می‌شود")
                print(f"✅ تعداد کل آگهی‌های پردازش شده: {processed_count}")
                print("💾 داده‌های استخراج شده در فایل اکسل ذخیره شده است")
                finalize_excel_save()
                break
            
            # اسکرول بهینه‌شده
            print("🔄 اسکرول برای بارگذاری آگهی‌های بیشتر...")
            
            # اسکرول مستقیم به انتها
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(1)  # کاهش انتظار
            
            # بررسی اینکه آیا آگهی جدیدی بارگذاری شده یا نه
            new_ad_links = getads()
            if len(new_ad_links) == len(ad_links):
                print("⚠️ آگهی جدیدی بارگذاری نشد، تلاش برای کلیک روی دکمه 'آگهی‌های بیشتر'...")
                
                # تلاش برای پیدا کردن و کلیک روی دکمه "آگهی‌های بیشتر"
                load_more_selectors = [
                    "#post-list-container-id > div.post-list__bottom-container-cac2f > div > button",
                    ".post-list__load-more-btn-be092",
                    "button[class*='load-more']",
                    "button[class*='post-list']"
                ]
                
                button_found = False
                for selector in load_more_selectors:
                    try:
                        load_more_button = driver.find_element(By.CSS_SELECTOR, selector)
                        if load_more_button and load_more_button.is_displayed():
                            print(f"🔘 دکمه بارگذاری بیشتر پیدا شد با selector: {selector}")
                            driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", load_more_button)
                            time.sleep(0.5)
                            load_more_button.click()
                            time.sleep(1)
                            print("✅ کلیک روی دکمه 'آگهی‌های بیشتر' موفق بود")
                            
                            # بررسی مجدد آگهی‌های جدید بعد از کلیک
                            time.sleep(1)
                            newer_ad_links = getads()
                            if len(newer_ad_links) > len(ad_links):
                                print("✅ آگهی‌های جدید بارگذاری شدند")
                                consecutive_pagination_errors = 0
                                button_found = True
                                break
                            else:
                                print("⚠️ بعد از کلیک روی دکمه، آگهی جدیدی بارگذاری نشد")
                    except Exception as e:
                        continue
                
                if not button_found:
                    print("❌ هیچ دکمه 'آگهی‌های بیشتر' پیدا نشد")
                
                # اگر دکمه پیدا شد و کار کرد، به حلقه اصلی برگرد
                if button_found:
                    continue
                
                print("🔄 تلاش برای بارگذاری آگهی‌های بیشتر...")
                
                # تلاش ساده برای پیدا کردن دکمه بارگذاری بیشتر
                try:
                    load_more_btn = driver.find_element(By.CSS_SELECTOR, "span.kt-button__ripple, button[class*='load-more']")
                    if load_more_btn and load_more_btn.is_displayed():
                        print("🔘 کلیک روی دکمه بارگذاری بیشتر...")
                        driver.execute_script("arguments[0].click();", load_more_btn)
                        time.sleep(1)
                        
                        # بررسی آگهی‌های جدید
                        newer_ad_links = getads()
                        if len(newer_ad_links) > len(ad_links):
                            print(f"✅ {len(newer_ad_links) - len(ad_links)} آگهی جدید بارگذاری شد")
                            consecutive_pagination_errors = 0
                        else:
                            consecutive_pagination_errors += 1
                    else:
                        consecutive_pagination_errors += 1
                except:
                    consecutive_pagination_errors += 1
                
                # اگر بعد از تمام اسکرول‌ها آگهی جدیدی پیدا نشد
                if len(getads()) == len(ad_links):
                    consecutive_pagination_errors += 1
                    print(f"⚠️ بعد از اسکرول، آگهی جدیدی پیدا نشد. خطای متوالی: {consecutive_pagination_errors}")
                    
                    # بررسی اینکه آیا به حداکثر خطاهای متوالی رسیده‌ایم
                    if consecutive_pagination_errors >= max_consecutive_pagination_errors:
                        print(f"❌ بعد از {max_consecutive_pagination_errors} خطای متوالی، بات متوقف می‌شود")
                        print(f"✅ تعداد کل آگهی‌های پردازش شده: {processed_count}")
                        print("💾 داده‌های استخراج شده در فایل اکسل ذخیره شده است")
                        finalize_excel_save()
                        break
                    
                    # صبر بیشتر قبل از تلاش مجدد
                    time.sleep(random.uniform(5, 10))
            else:
                # اگر آگهی جدید بارگذاری شد، شمارنده را ریست کنیم
                consecutive_pagination_errors = 0
                
    except Exception as e:
        print(f"Error getting ads: {e}")
        
        # بررسی اینکه آیا خطا مربوط به crash driver است
        error_str = str(e).lower()
        
        # کلمات کلیدی برای تشخیص crash واقعی driver
        critical_crash_keywords = [
            'invalid session id', 'session not created', 'chrome not reachable',
            'disconnected', 'crashed', 'no such session', 'session deleted',
            'connection refused', 'Connection refused', 'chrome has crashed',
            'httpconnectionpool', 'HTTPConnectionPool'
        ]
        
        # بررسی driver availability
        driver_is_alive = False
        try:
            # تست ساده برای بررسی زنده بودن driver
            driver.current_url
            driver_is_alive = True
        except:
            driver_is_alive = False
        
        # فقط در صورت crash واقعی driver restart کن
        if not driver_is_alive or any(keyword in error_str for keyword in critical_crash_keywords):
            print("🚨 تشخیص crash واقعی در Chrome driver")
            restart_attempts += 1
            
            if restart_attempts <= max_restart_attempts:
                print(f"تلاش {restart_attempts} از {max_restart_attempts} برای راه‌اندازی مجدد...")
                if restart_driver():
                    print("✅ Driver با موفقیت راه‌اندازی شد، ادامه پردازش...")
                    # ریست کردن شمارنده restart بعد از موفقیت
                    restart_attempts = 0
                    continue
                else:
                    print(f"❌ خطا در راه‌اندازی مجدد، تلاش {restart_attempts}")
                    time.sleep(random.uniform(10, 20))
                    continue
            else:
                print(f"❌ حداکثر تعداد تلاش برای راه‌اندازی مجدد ({max_restart_attempts}) به پایان رسید")
                finalize_excel_save()
                break
        else:
            # خطای عادی، فقط صبر کن و ادامه بده (بدون restart)
            print("⚠️ خطای عادی (driver سالم است)، ادامه پردازش...")
            time.sleep(random.uniform(3, 8))
            continue

# ذخیره نهایی داده‌های باقی‌مانده
finalize_excel_save()
print("✅ تمام داده‌های باقی‌مانده ذخیره شدند")

# بستن درایور
try:
    driver.quit()
except:
    pass

print(f"🎉 پردازش کامل شد! تعداد کل آگهی‌های پردازش شده: {processed_count}")
print("📁 فایل‌های اکسل در پوشه پروژه ذخیره شده‌اند")
